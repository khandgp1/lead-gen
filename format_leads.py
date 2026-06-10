import openpyxl
import copy
import glob
from openpyxl.worksheet.hyperlink import Hyperlink

# ─── Target schema ────────────────────────────────────────────────────────────
TARGET_HEADERS = [
    'Business Name', 'Location', 'Website / Online Presence',
    'Phone', 'Email', 'Opportunity Reason', 'Sources Verified', 'Validated'
]

TARGET_WIDTHS = {
    'A': 28.0,
    'B': 16.0,
    'C': 35.0,
    'D': 16.0,
    'E': 29.6640625,
    'F': 55.0,
    'G': 38.0,
    'H': 32.0,
}

# ─── Variant column mappings (target_col_letter → source_col_index, 1-based) ──
# None means the target column should be left blank.

# Variant A: 10-col, header row 3, Rank(A)+WebsiteStatus(D)+Validated(J)
# B→A, C→B, E→C, F→D, G→E, H→F, I→G, J→H
VARIANT_A_MAP = {
    'A': 2, 'B': 3, 'C': 5, 'D': 6,
    'E': 7, 'F': 8, 'G': 9, 'H': 10,
}

# Variant A2: 9-col, header row 3, Rank(A)+WebsiteStatus(D), no State License Board, no Validated
# B→A, C→B, E→C, F→D, G→E, H→F, I→G, H=blank
VARIANT_A2_MAP = {
    'A': 2, 'B': 3, 'C': 5, 'D': 6,
    'E': 7, 'F': 8, 'G': 9, 'H': None,
}

# Variant B/C: 10-col, Rank(A)+WebsiteStatus(D)+StateLicenseBoard(H), no Validated
# B→A, C→B, E→C, F→D, G→E, I→F, J→G, H=blank
VARIANT_BC_MAP = {
    'A': 2, 'B': 3, 'C': 5, 'D': 6,
    'E': 7, 'F': 9, 'G': 10, 'H': None,
}


def copy_cell_style(src_cell, dst_cell):
    """Copy all style properties from src to dst safely."""
    if src_cell.font:
        dst_cell.font = copy.copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy.copy(src_cell.fill)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.border:
        dst_cell.border = copy.copy(src_cell.border)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format


def detect_variant(ws):
    """
    Returns a tuple: (variant, header_row)
    - 'standard'  : already 8 cols with correct headers (skip)
    - 'A'         : 10-col, header row 3, has Validated
    - 'A2'        : 9-col,  header row 3, no State License Board, no Validated
    - 'B'         : 10-col, header row 3, no Validated (has State License Board)
    - 'C'         : 10-col, header row 4, no Validated (blank row 3, State License Board)
    """
    row3 = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    row4 = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]

    # Standard: row 3 matches target exactly
    if row3[:8] == TARGET_HEADERS and ws.max_column == 8:
        return 'standard', 3

    # Row 3 starts with 'Rank'
    if row3[0] == 'Rank':
        if 'Validated' in row3:
            return 'A', 3
        elif ws.max_column == 9:
            return 'A2', 3   # file 04: 9-col, no Validated, no State License Board
        else:
            return 'B', 3

    # Variant C: blank row 3, header in row 4
    if row3[0] is None and row4[0] == 'Rank':
        return 'C', 4

    return 'unknown', None


def format_file(filepath):
    print(f"\nProcessing: {filepath}")

    wb_orig = openpyxl.load_workbook(filepath)
    ws_orig = wb_orig.active  # Use active sheet regardless of name

    variant, header_row = detect_variant(ws_orig)

    if variant == 'standard':
        print(f"  ✓ Already standard — skipping.")
        wb_orig.close()
        return

    if variant == 'unknown':
        print(f"  ⚠ Unknown variant — skipping.")
        wb_orig.close()
        return

    print(f"  Detected variant {variant} (header row {header_row})")

    col_map = VARIANT_A_MAP if variant == 'A' else (VARIANT_A2_MAP if variant == 'A2' else VARIANT_BC_MAP)

    # Data in source starts at header_row; for variant C header is row 4, data row 5
    src_data_start = header_row  # header row itself is first "data" row to copy

    # ── Build new workbook ────────────────────────────────────────────────────
    wb_new = openpyxl.Workbook()
    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    ws_new = wb_new.create_sheet(title="Ranked Leads")

    # Copy sheet view settings
    ws_new.sheet_view.zoomScale = ws_orig.sheet_view.zoomScale or 110
    ws_new.sheet_view.tabSelected = ws_orig.sheet_view.tabSelected
    ws_new.sheet_view.showGridLines = ws_orig.sheet_view.showGridLines

    # Set standard column widths
    for col_letter, w in TARGET_WIDTHS.items():
        ws_new.column_dimensions[col_letter].width = w

    # ── Title rows 1 & 2 ─────────────────────────────────────────────────────
    # For variant C the title rows are at rows 1 & 2 in the source too.
    # Row 3 in source is blank (skip it). Header in source is at row 4 → maps to row 3 in dest.
    for r in [1, 2]:
        # Copy row height from original
        if r in ws_orig.row_dimensions and ws_orig.row_dimensions[r].height is not None:
            ws_new.row_dimensions[r].height = ws_orig.row_dimensions[r].height

        src_cell = ws_orig.cell(row=r, column=1)  # A1 / A2 holds the title/subtitle
        dst_cell = ws_new.cell(row=r, column=1)
        dst_cell.value = src_cell.value
        copy_cell_style(src_cell, dst_cell)

    ws_new.merge_cells("A1:H1")
    ws_new.merge_cells("A2:H2")

    # ── Header + data rows ────────────────────────────────────────────────────
    # src rows from header_row to max_row → dst rows from 3 onwards
    dst_row = 3
    for src_row in range(header_row, ws_orig.max_row + 1):
        # Copy row height; for variant C the row heights shift by 1
        src_dim = ws_orig.row_dimensions.get(src_row)
        if src_dim and src_dim.height is not None:
            ws_new.row_dimensions[dst_row].height = src_dim.height

        for target_col, src_col_idx in col_map.items():
            dst_cell = ws_new[f"{target_col}{dst_row}"]

            if src_col_idx is None:
                # Blank target column (Validated for variant B/C)
                continue

            src_cell = ws_orig.cell(row=src_row, column=src_col_idx)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

            if src_cell.hyperlink:
                dst_cell.hyperlink = Hyperlink(
                    target=src_cell.hyperlink.target,
                    ref=f"{target_col}{dst_row}"
                )

        dst_row += 1

    # ── Normalize column C header to standard name ────────────────────────────
    # Source files may use 'Facebook / Nextdoor', 'Facebook / Online Presence', etc.
    # Force it to the canonical target header name.
    ws_new['C3'].value = 'Website / Online Presence'

    wb_new.save(filepath)
    print(f"  ✓ Saved ({dst_row - 3} data+header rows written).")


def main():
    files = sorted(glob.glob("input/*.xlsx"))
    print(f"Found {len(files)} xlsx files in input/")

    for filepath in files:
        format_file(filepath)

    print("\nAll files processed.")


if __name__ == "__main__":
    main()
