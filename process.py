import os
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.worksheet.hyperlink import Hyperlink

def copy_font(src_font):
    if not src_font:
        return None
    color = None
    if src_font.color:
        if src_font.color.type == 'theme':
            color = Color(theme=src_font.color.theme, tint=src_font.color.tint)
        elif src_font.color.rgb:
            color = Color(rgb=src_font.color.rgb)
    return Font(
        name=src_font.name,
        size=src_font.size,
        bold=src_font.bold,
        italic=src_font.italic,
        underline=src_font.underline,
        color=color
    )


def extract_leads_from_file(filepath):
    """
    Open a single standardized input file and return a list of lead dicts.
    Returns an empty list if the file cannot be read or has no 'Ranked Leads' sheet.
    """
    # Columns mapping: Target letter -> Standard col number (1-based, post-format_leads)
    col_mapping = {
        'A': 1,  # Business Name
        'B': 2,  # Location
        'C': 3,  # Website / Online Presence
        'D': 4,  # Phone
        'E': 5,  # Email
        'F': 7,  # Sources Verified (col G in standard schema)
    }

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        print(f"  ⚠ Could not open {filepath}: {e} — skipping.")
        return []

    if "Ranked Leads" not in wb.sheetnames:
        print(f"  ⚠ No 'Ranked Leads' sheet in {filepath} — skipping.")
        wb.close()
        return []

    ws = wb["Ranked Leads"]
    leads = []

    for r in range(4, ws.max_row + 1):
        email_cell = ws.cell(row=r, column=5)
        email_val = email_cell.value

        # Filter: skip rows with no valid email
        if not email_val or str(email_val).strip().lower() in ["", "no email found", "not found — check fb/nd"]:
            continue

        row_data = {}
        for target_col, orig_col_idx in col_mapping.items():
            cell = ws.cell(row=r, column=orig_col_idx)
            val = cell.value if target_col != 'F' else None  # Sources Verified left blank
            hyperlink_target = cell.hyperlink.target if cell.hyperlink else None
            font_obj = copy_font(cell.font)

            row_data[target_col] = {
                "value": val,
                "hyperlink_target": hyperlink_target,
                "font": font_obj
            }
        leads.append(row_data)

    wb.close()
    return leads


def main():
    input_dir = "input"
    output_dir = "output"
    output_filename = "Roofing_Outreach_Leads.xlsx"
    output_filepath = os.path.join(output_dir, output_filename)

    # ── Gather all input files ────────────────────────────────────────────────
    all_files = sorted(glob.glob(os.path.join(input_dir, "*.xlsx")))
    print(f"Found {len(all_files)} xlsx file(s) in '{input_dir}/'")

    # ── Extract leads from each file ──────────────────────────────────────────
    all_leads = []
    for filepath in all_files:
        file_leads = extract_leads_from_file(filepath)
        print(f"  {os.path.basename(filepath)}: {len(file_leads)} lead(s) with valid email")
        all_leads.extend(file_leads)

    print(f"\nTotal leads extracted: {len(all_leads)}")

    if not all_leads:
        print("No leads to write — output file not created.")
        return

    # ── Build output workbook ─────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    default_sheet = wb_out.active
    wb_out.remove(default_sheet)

    ws_out = wb_out.create_sheet(title='Leads')
    ws_out.sheet_view.zoomScale = 110
    ws_out.sheet_view.tabSelected = True
    ws_out.sheet_view.showGridLines = None

    # Column widths
    widths = {
        'A': 28.0,
        'B': 16.0,
        'C': 35.0,
        'D': 16.0,
        'E': 29.6640625,
        'F': 38.0
    }
    for col_letter, w in widths.items():
        ws_out.column_dimensions[col_letter].width = w

    # Row heights
    ws_out.row_dimensions[1].height = 30.0
    ws_out.row_dimensions[2].height = 16.0
    ws_out.row_dimensions[3].height = 27.75

    # Merge rows 1 & 2
    ws_out.merge_cells('A1:F1')
    ws_out.merge_cells('A2:F2')

    # Title row 1
    a1 = ws_out['A1']
    a1.font = Font(name='Arial', size=14.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    a1.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF1F3864'))
    a1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=None)

    # Metadata row 2
    a2 = ws_out['A2']
    a2.font = Font(name='Arial', size=9.0, bold=False, italic=True, underline=None, color=Color(rgb='FF808080'))
    a2.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFF2F2F2'))
    a2.alignment = Alignment(horizontal='center', vertical=None, wrap_text=None)

    # Style empty cells in rows 1 & 2
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_out[f"{col}1"].font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
        ws_out[f"{col}2"].font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))

    # Column header row 3
    headers = {
        'A': 'Business Name',
        'B': 'Location',
        'C': 'Website / Online Presence',
        'D': 'Phone',
        'E': 'Email',
        'F': 'Sources Verified'
    }
    for col, val in headers.items():
        cell = ws_out[f"{col}3"]
        cell.value = val
        cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
        cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
        wrap_hdr = False if col == 'C' else True
        shrink_hdr = True if col == 'C' else None
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap_hdr, shrink_to_fit=shrink_hdr)
        cell.border = Border(
            left=Side(style='thin', color=Color(rgb='FFBFBFBF')),
            right=Side(style='thin', color=Color(rgb='FFBFBFBF')),
            top=Side(style='medium', color=Color(rgb='FF2E75B6')),
            bottom=Side(style='thin', color=Color(rgb='FFBFBFBF'))
        )

    # Data rows (start at row 4)
    border_style = Border(
        left=Side(style='thin', color=Color(rgb='FFBFBFBF')),
        right=Side(style='thin', color=Color(rgb='FFBFBFBF')),
        top=Side(style='thin', color=Color(rgb='FFBFBFBF')),
        bottom=Side(style='thin', color=Color(rgb='FFBFBFBF'))
    )

    for idx, row_data in enumerate(all_leads):
        r = 4 + idx
        fill_color = 'FFE2EFDA' if r % 2 == 0 else 'FFFFFFFF'
        row_fill = PatternFill(fill_type='solid', fgColor=Color(rgb=fill_color))

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws_out[f"{col_letter}{r}"]
            cell.fill = row_fill
            cell.border = border_style

            cell_info = row_data[col_letter]
            val = cell_info["value"]
            target = cell_info["hyperlink_target"]
            f_style = cell_info["font"]

            cell.value = val
            cell.font = f_style
            wrap_data = False if col_letter == 'C' else True
            shrink_data = True if col_letter == 'C' else None
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=wrap_data, shrink_to_fit=shrink_data)

            if target:
                cell.hyperlink = Hyperlink(target=target, ref=f"{col_letter}{r}")

    # ── Save output ───────────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    wb_out.save(output_filepath)
    print(f"\nOutput saved to: {output_filepath}  ({len(all_leads)} rows)")


if __name__ == "__main__":
    main()
