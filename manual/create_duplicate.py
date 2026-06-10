import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.worksheet.hyperlink import Hyperlink

def create_duplicate():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheet Ranked Leads
    ws = wb.create_sheet(title='Ranked Leads')
    ws.sheet_view.zoomScale = 110
    ws.sheet_view.tabSelected = True
    ws.sheet_view.showGridLines = None

    # Column widths
    ws.column_dimensions['A'].width = 28.0
    ws.column_dimensions['B'].width = 16.0
    ws.column_dimensions['C'].width = 35.0
    ws.column_dimensions['D'].width = 16.0
    ws.column_dimensions['E'].width = 29.6640625
    ws.column_dimensions['F'].width = 38.0
    # Row heights
    ws.row_dimensions[1].height = 30.0
    ws.row_dimensions[2].height = 16.0
    ws.row_dimensions[3].height = 27.75
    ws.row_dimensions[4].height = 72.0
    ws.row_dimensions[5].height = 72.0
    ws.row_dimensions[6].height = 72.0
    ws.row_dimensions[7].height = 72.0
    ws.row_dimensions[8].height = 72.0
    ws.row_dimensions[9].height = 72.0
    ws.row_dimensions[10].height = 72.0
    ws.row_dimensions[11].height = 72.0
    ws.row_dimensions[12].height = 72.0
    ws.row_dimensions[13].height = 72.0
    ws.row_dimensions[14].height = 72.0
    ws.row_dimensions[15].height = 72.0
    ws.row_dimensions[16].height = 72.0
    ws.row_dimensions[17].height = 72.0
    ws.row_dimensions[18].height = 72.0
    # Merged ranges
    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')
    # Cells
    cell = ws['A1']
    cell.font = Font(name='Arial', size=14.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF1F3864'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=None)
    cell = ws['B1']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['C1']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['D1']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['E1']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['F1']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['A2']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=True, underline=None, color=Color(rgb='FF808080'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFF2F2F2'))
    cell.alignment = Alignment(horizontal='center', vertical=None, wrap_text=None)
    cell = ws['B2']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['C2']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['D2']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['E2']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['F2']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline=None, color=Color(theme=1, tint=0.0))
    cell = ws['A3']
    cell.value = 'Business Name'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B3']
    cell.value = 'Location'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C3']
    cell.value = 'Website / Online Presence'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D3']
    cell.value = 'Phone'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E3']
    cell.value = 'Email'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F3']
    cell.value = 'Sources Verified'
    cell.font = Font(name='Arial', size=10.0, bold=True, italic=False, underline=None, color=Color(rgb='FFFFFFFF'))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FF2E75B6'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='medium', color=Color(rgb='FF2E75B6')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A4']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B4']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C4']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D4']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E4']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F4']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A5']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B5']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C5']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D5']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E5']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F5']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A6']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B6']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C6']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D6']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E6']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F6']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F7']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A8']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B8']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C8']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D8']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E8']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F8']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A9']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B9']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C9']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D9']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E9']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F9']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F10']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F11']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A12']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B12']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C12']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D12']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E12']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F12']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F13']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A14']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B14']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C14']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D14']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E14']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F14']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A15']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B15']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C15']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D15']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E15']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F15']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A16']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B16']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C16']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D16']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E16']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F16']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A17']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B17']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C17']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D17']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E17']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F17']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFFFFFFF'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['A18']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['B18']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['C18']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['D18']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['E18']
    cell.font = Font(name='Calibri', size=11.0, bold=False, italic=False, underline='single', color=Color(theme=10, tint=0.0))
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))
    cell = ws['F18']
    cell.font = Font(name='Arial', size=9.0, bold=False, italic=False, underline=None)
    cell.fill = PatternFill(fill_type='solid', fgColor=Color(rgb='FFE2EFDA'))
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = Border(left=Side(style='thin', color=Color(rgb='FFBFBFBF')), right=Side(style='thin', color=Color(rgb='FFBFBFBF')), top=Side(style='thin', color=Color(rgb='FFBFBFBF')), bottom=Side(style='thin', color=Color(rgb='FFBFBFBF')))

    wb.save('/Users/khandpv1/Desktop/.AntiGrav/lead-gen/output/Roofing_Outreach_Leads_Template.xlsx')
    print('Duplicate saved to output/Roofing_Outreach_Leads_Template.xlsx')

if __name__ == '__main__':
    create_duplicate()