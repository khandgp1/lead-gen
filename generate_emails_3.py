import os
import openpyxl

# ── Template ──────────────────────────────────────────────────────────────────

SUBJECT_TEMPLATE = "Website for {business_name}"

BODY_TEMPLATE = """\
Hi,

Those thousands of {city} searches are going to roofers with websites.

Want to see one for {business_name}?"""

# ── HTML builder ──────────────────────────────────────────────────────────────

def html_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;")
    )


def build_html(leads: list[dict]) -> str:
    cards_html = ""
    for i, lead in enumerate(leads, start=1):
        name    = html_escape(lead["name"])
        email   = html_escape(lead["email"])
        website = html_escape(lead["website"])
        subject = html_escape(lead["subject"])
        body    = html_escape(lead["body"])

        # clipboard payload stored in a data attribute (JSON-safe via html_escape)
        copy_text = html_escape(
            f"To: {lead['email']}\nSubject: {lead['subject']}\n\n{lead['body']}"
        )

        cards_html += f"""
    <div class="card" id="card-{i}">
      <div class="card-header">
        <span class="badge">#{i}</span>
        <span class="biz-name">{name}</span>
        <span class="card-count">{i} of {len(leads)}</span>
      </div>

      <div class="field-row">
        <span class="label">To</span>
        <span class="value email-addr">{email}</span>
      </div>

      <div class="field-row">
        <span class="label">Website</span>
        <span class="value website-val">{website}</span>
      </div>

      <div class="field-row">
        <span class="label">Subject</span>
        <span class="value">{subject}</span>
      </div>

      <div class="field-row body-row">
        <span class="label">Body</span>
        <pre class="body-text">{body}</pre>
      </div>

      <button
        class="copy-btn"
        data-copy="{copy_text}"
        onclick="copyAll(this)"
      >
        Copy All
      </button>
    </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Roofing Outreach Emails - Sequence 3 ({len(leads)} leads)</title>
  <link rel="preconnect" href="https://fonts.googleapis.com" />
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet" />
  <style>
    /* ── Reset & tokens ─────────────────────────────────── */
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

    :root {{
      --bg:        #0d1117;
      --surface:   #161b22;
      --surface2:  #21262d;
      --border:    #30363d;
      --accent:    #2563eb;
      --accent-h:  #3b82f6;
      --accent-glow: rgba(37,99,235,0.25);
      --success:   #22c55e;
      --text:      #e6edf3;
      --text-muted:#8b949e;
      --label-bg:  #1c2230;
      --radius:    12px;
      --radius-sm: 6px;
    }}

    body {{
      font-family: 'Inter', sans-serif;
      background: var(--bg);
      color: var(--text);
      min-height: 100vh;
      padding: 0 0 80px;
    }}

    /* ── Sticky header ──────────────────────────────────── */
    .site-header {{
      position: sticky;
      top: 0;
      z-index: 100;
      background: rgba(13,17,23,0.85);
      backdrop-filter: blur(12px);
      border-bottom: 1px solid var(--border);
      padding: 18px 32px;
      display: flex;
      align-items: center;
      gap: 20px;
    }}

    .header-title {{
      font-size: 18px;
      font-weight: 700;
      letter-spacing: -0.3px;
      white-space: nowrap;
    }}

    .header-title span {{
      background: linear-gradient(135deg, #60a5fa, #a78bfa);
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
    }}

    .lead-count {{
      font-size: 13px;
      font-weight: 500;
      color: var(--text-muted);
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: 20px;
      padding: 4px 12px;
      white-space: nowrap;
    }}

    .search-wrap {{
      flex: 1;
      max-width: 380px;
      margin-left: auto;
      position: relative;
    }}

    .search-wrap svg {{
      position: absolute;
      left: 12px;
      top: 50%;
      transform: translateY(-50%);
      color: var(--text-muted);
      pointer-events: none;
    }}

    #search {{
      width: 100%;
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: var(--radius-sm);
      color: var(--text);
      font-family: inherit;
      font-size: 14px;
      padding: 8px 12px 8px 36px;
      outline: none;
      transition: border-color .2s, box-shadow .2s;
    }}

    #search:focus {{
      border-color: var(--accent);
      box-shadow: 0 0 0 3px var(--accent-glow);
    }}

    #search::placeholder {{ color: var(--text-muted); }}

    /* ── Grid ───────────────────────────────────────────── */
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(520px, 1fr));
      gap: 20px;
      padding: 32px;
      max-width: 1400px;
      margin: 0 auto;
    }}

    /* ── Card ───────────────────────────────────────────── */
    .card {{
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 22px 24px 20px;
      display: flex;
      flex-direction: column;
      gap: 14px;
      transition: border-color .2s, box-shadow .2s, transform .15s;
    }}

    .card:hover {{
      border-color: #3d444d;
      box-shadow: 0 8px 32px rgba(0,0,0,0.35);
      transform: translateY(-2px);
    }}

    /* ── Card header ────────────────────────────────────── */
    .card-header {{
      display: flex;
      align-items: center;
      gap: 10px;
    }}

    .badge {{
      font-size: 11px;
      font-weight: 700;
      background: linear-gradient(135deg, var(--accent), #7c3aed);
      border-radius: 20px;
      padding: 3px 9px;
      color: #fff;
      letter-spacing: .3px;
      flex-shrink: 0;
    }}

    .biz-name {{
      font-size: 15px;
      font-weight: 700;
      color: var(--text);
      flex: 1;
      letter-spacing: -0.2px;
    }}

    .card-count {{
      font-size: 11px;
      color: var(--text-muted);
      white-space: nowrap;
    }}

    /* ── Field rows ─────────────────────────────────────── */
    .field-row {{
      display: flex;
      gap: 10px;
      align-items: baseline;
    }}

    .label {{
      font-size: 11px;
      font-weight: 600;
      text-transform: uppercase;
      letter-spacing: .6px;
      color: var(--text-muted);
      background: var(--label-bg);
      border: 1px solid var(--border);
      border-radius: 4px;
      padding: 2px 7px;
      white-space: nowrap;
      flex-shrink: 0;
    }}

    .value {{
      font-size: 14px;
      color: var(--text);
      word-break: break-all;
    }}

    .email-addr {{
      color: #60a5fa;
      font-weight: 500;
    }}

    /* ── Body field ─────────────────────────────────────── */
    .body-row {{
      align-items: flex-start;
    }}

    .body-text {{
      font-family: 'Inter', sans-serif;
      font-size: 13.5px;
      line-height: 1.7;
      color: #c9d1d9;
      white-space: pre-wrap;
      word-break: break-word;
      background: var(--surface2);
      border: 1px solid var(--border);
      border-radius: var(--radius-sm);
      padding: 14px 16px;
      flex: 1;
    }}

    /* ── Copy button ────────────────────────────────────── */
    .copy-btn {{
      align-self: flex-end;
      background: var(--accent);
      color: #fff;
      font-family: inherit;
      font-size: 13px;
      font-weight: 600;
      border: none;
      border-radius: var(--radius-sm);
      padding: 9px 20px;
      cursor: pointer;
      transition: background .2s, box-shadow .2s, transform .1s;
      letter-spacing: .2px;
    }}

    .copy-btn:hover {{
      background: var(--accent-h);
      box-shadow: 0 4px 16px var(--accent-glow);
    }}

    .copy-btn:active {{ transform: scale(.97); }}

    .copy-btn.copied {{
      background: var(--success);
      box-shadow: 0 4px 16px rgba(34,197,94,.25);
    }}

    /* ── Empty state ────────────────────────────────────── */
    .empty {{
      display: none;
      grid-column: 1 / -1;
      text-align: center;
      padding: 80px 20px;
      color: var(--text-muted);
    }}

    .empty svg {{ margin-bottom: 16px; opacity: .4; }}
    .empty p {{ font-size: 15px; }}

    /* ── Scrollbar ──────────────────────────────────────── */
    ::-webkit-scrollbar {{ width: 8px; }}
    ::-webkit-scrollbar-track {{ background: var(--bg); }}
    ::-webkit-scrollbar-thumb {{ background: var(--border); border-radius: 4px; }}
    ::-webkit-scrollbar-thumb:hover {{ background: #484f58; }}
  </style>
</head>
<body>

  <!-- Sticky header -->
  <header class="site-header">
    <div class="header-title">Roofing Outreach &mdash; <span>Sequence 3 Viewer</span></div>
    <div class="lead-count">{len(leads)} emails</div>
    <div class="search-wrap">
      <svg width="15" height="15" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
        <circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/>
      </svg>
      <input id="search" type="text" placeholder="Search by business name or email…" oninput="filterCards(this.value)" />
    </div>
  </header>

  <!-- Card grid -->
  <div class="grid" id="grid">
{cards_html}
    <div class="empty" id="empty-state">
      <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5">
        <circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/>
      </svg>
      <p>No emails match your search.</p>
    </div>
  </div>

  <script>
    function copyAll(btn) {{
      const text = btn.getAttribute('data-copy')
        .replace(/&amp;/g, '&')
        .replace(/&lt;/g, '<')
        .replace(/&gt;/g, '>')
        .replace(/&quot;/g, '"')
        .replace(/&#39;/g, "'");

      navigator.clipboard.writeText(text).then(() => {{
        btn.textContent = 'Copied ✓';
        btn.classList.add('copied');
        setTimeout(() => {{
          btn.textContent = 'Copy All';
          btn.classList.remove('copied');
        }}, 2000);
      }}).catch(() => {{
        // Fallback for non-HTTPS
        const ta = document.createElement('textarea');
        ta.value = text;
        ta.style.position = 'fixed';
        ta.style.opacity = '0';
        document.body.appendChild(ta);
        ta.select();
        document.execCommand('copy');
        document.body.removeChild(ta);
        btn.textContent = 'Copied ✓';
        btn.classList.add('copied');
        setTimeout(() => {{
          btn.textContent = 'Copy All';
          btn.classList.remove('copied');
        }}, 2000);
      }});
    }}

    function filterCards(query) {{
      const q = query.toLowerCase().trim();
      const cards = document.querySelectorAll('.card');
      let visible = 0;
      cards.forEach(card => {{
        const text = card.textContent.toLowerCase();
        const show = !q || text.includes(q);
        card.style.display = show ? '' : 'none';
        if (show) visible++;
      }});
      document.getElementById('empty-state').style.display = visible === 0 ? 'block' : 'none';
    }}
  </script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    input_path  = os.path.join("manual", "VALID_2_Roofing_Outreach_Leads.xlsx")
    output_dir  = "output"
    output_path = os.path.join(output_dir, "sequence_3_emails.html")

    print(f"Reading leads from: {input_path}")
    wb = openpyxl.load_workbook(input_path)

    if "Leads" not in wb.sheetnames:
        print("ERROR: 'Leads' sheet not found in workbook.")
        return

    ws = wb["Leads"]
    leads = []
    skipped = 0

    for r in range(3, ws.max_row + 1):
        raw_name     = ws.cell(row=r, column=1).value
        raw_location = ws.cell(row=r, column=2).value
        raw_website  = ws.cell(row=r, column=3).value
        raw_email    = ws.cell(row=r, column=6).value

        name     = str(raw_name).strip()     if raw_name     else ""
        location = str(raw_location).strip() if raw_location else ""
        website  = str(raw_website).strip()  if raw_website  else "—"
        email    = str(raw_email).strip()    if raw_email    else ""

        if not name or not email:
            skipped += 1
            continue

        # Use the full location (e.g. Joplin, MO) including the state
        city = location if location else ""

        subject = SUBJECT_TEMPLATE.format(business_name=name)
        body    = BODY_TEMPLATE.format(business_name=name, city=city)

        leads.append({
            "name":    name,
            "email":   email,
            "website": website,
            "subject": subject,
            "body":    body,
        })

    wb.close()

    print(f"  Leads loaded : {len(leads)}")
    if skipped:
        print(f"  Rows skipped : {skipped} (missing name or email)")

    if not leads:
        print("No leads to generate — aborting.")
        return

    html = build_html(leads)
    os.makedirs(output_dir, exist_ok=True)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n✅  Output written to: {output_path}")
    print(f"    Open in your browser to view & copy emails.")


if __name__ == "__main__":
    main()
